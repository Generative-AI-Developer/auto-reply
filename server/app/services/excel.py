"""Parse a bulk-upload .xlsx into RequestCreate rows.

Expected headers (case-insensitive; extra columns ignored):
  Request Number | Numbers | Request Type | Duration Days | Case Officer | Justification
`Numbers` (Mobile/CNIC/IMEI No) may hold several values separated by comma /
semicolon / newline. `Request Number` is required and is user-supplied
(unique per requester) - see routers/requests.py::_create_request for the
merge-into-existing-request behavior when it repeats.

request_date is not read from the sheet: every request is stamped with the
system's current date at creation time (see routers/requests.py).
"""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook

from ..schemas import RequestCreate

_HEADER_ALIASES = {
    "request number": "request_number",
    "request no": "request_number",
    "numbers": "numbers",
    "number": "numbers",
    "mobile": "numbers",
    "nic": "numbers",
    "imei": "numbers",
    "mobile/cnic/imei no": "numbers",
    "request type": "request_type",
    "type": "request_type",
    "duration days": "duration_days",
    "duration": "duration_days",
    "days": "duration_days",
    "case officer": "case_officer",
    "officer": "case_officer",
    "justification": "justification",
    "reason": "justification",
}

_SPLIT_RE = re.compile(r"[,;\n]+")


def parse_excel(data: bytes) -> tuple[list[RequestCreate], list[str]]:
    rows: list[RequestCreate] = []
    errors: list[str] = []

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        wb.close()
        return rows, ["Empty sheet"]

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_cells):
        if cell is None:
            continue
        key = _HEADER_ALIASES.get(str(cell).strip().lower())
        if key:
            col_map[idx] = key

    if "numbers" not in col_map.values():
        wb.close()
        return rows, ["Missing required 'Mobile/CNIC/IMEI No' column"]
    if "request_number" not in col_map.values():
        wb.close()
        return rows, ["Missing required 'Request Number' column"]

    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or c == "" for c in row):
            continue
        fields: dict = {}
        for idx, key in col_map.items():
            fields[key] = row[idx] if idx < len(row) else None

        request_number = str(fields.get("request_number") or "").strip()
        if not request_number:
            errors.append(f"Row {row_no}: no request number")
            continue

        raw_numbers = fields.get("numbers")
        numbers = (
            [n.strip() for n in _SPLIT_RE.split(str(raw_numbers)) if n.strip()]
            if raw_numbers is not None
            else []
        )
        if not numbers:
            errors.append(f"Row {row_no}: no numbers")
            continue

        duration = fields.get("duration_days")
        try:
            duration_days = int(duration) if duration not in (None, "") else None
        except (TypeError, ValueError):
            duration_days = None

        rows.append(
            RequestCreate(
                request_number=request_number,
                numbers=numbers,
                request_type=str(fields.get("request_type") or ""),
                duration_days=duration_days,
                case_officer=str(fields.get("case_officer") or ""),
                justification=str(fields.get("justification") or ""),
            )
        )

    wb.close()
    return rows, errors
