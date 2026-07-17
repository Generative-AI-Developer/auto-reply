"""Extract candidate identifier numbers and a date from an incoming file.

Strategy (per plan): look at the filename first, then fall back to parsing the
file's content. Numbers are normalized to digits-only so `0300-123 4567` and
`03001234567` compare equal.
"""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from pathlib import Path

# A "number" worth matching has at least this many digits (avoids matching noise).
MIN_DIGITS = 7

_NUMBER_RE = re.compile(r"\d[\d\s\-]{%d,}\d" % (MIN_DIGITS - 2))
_DIGITS_RE = re.compile(r"\d+")

# Common date formats, most specific first. `(?<!\d)`/`(?!\d)` (instead of \b)
# so an underscore separator like `_2026-07-14` is still treated as a boundary.
_DATE_PATTERNS = [
    (re.compile(r"(?<!\d)(\d{4})[-_/.](\d{1,2})[-_/.](\d{1,2})(?!\d)"), "ymd"),
    (re.compile(r"(?<!\d)(\d{1,2})[-_/.](\d{1,2})[-_/.](\d{4})(?!\d)"), "dmy"),
    (re.compile(r"(?<!\d)(\d{4})(\d{2})(\d{2})(?!\d)"), "ymd"),  # 20260714
]


def normalize_number(raw: str) -> str:
    return "".join(_DIGITS_RE.findall(raw))


def extract_numbers_from_text(text: str) -> set[str]:
    numbers: set[str] = set()
    for m in _NUMBER_RE.finditer(text or ""):
        norm = normalize_number(m.group())
        if len(norm) >= MIN_DIGITS:
            numbers.add(norm)
    return numbers


def extract_date_from_text(text: str) -> date | None:
    for pattern, order in _DATE_PATTERNS:
        m = pattern.search(text or "")
        if not m:
            continue
        a, b, c = m.groups()
        try:
            if order == "ymd":
                return date(int(a), int(b), int(c))
            return date(int(c), int(b), int(a))
        except ValueError:
            continue
    return None


def _strip_dates(text: str) -> str:
    """Remove date substrings so their digits aren't picked up as identifier numbers."""
    for pattern, _ in _DATE_PATTERNS:
        text = pattern.sub(" ", text or "")
    return text


def _read_content(path: Path) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix in {".txt", ".csv"}:
            return path.read_text(encoding="utf-8", errors="ignore")
        if suffix == ".pdf":
            from pypdf import PdfReader

            reader = PdfReader(str(path))
            return "\n".join((page.extract_text() or "") for page in reader.pages)
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True, data_only=True)
            parts: list[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    parts.extend(str(c) for c in row if c is not None)
            wb.close()
            return " ".join(parts)
    except Exception:
        # Parsing best-effort; a bad/locked file simply yields no content matches.
        return ""
    return ""


def _a_number_column_values(path: Path) -> set[str] | None:
    """CDR-style CSVs carry many digit-heavy columns (IMEI, IMSI, B Number,
    Cell Id, lat/long...) that are noise for matching - only the "A Number"
    column identifies the subject of the file. Returns None (not this
    format) so the caller falls back to generic full-text extraction."""
    if path.suffix.lower() != ".csv":
        return None
    try:
        with path.open(newline="", encoding="utf-8-sig", errors="ignore") as f:
            reader = csv.reader(f)
            header = next(reader)
            if "A Number" not in header:
                return None
            idx = header.index("A Number")
            numbers: set[str] = set()
            for row in reader:
                if len(row) <= idx:
                    continue
                norm = normalize_number(row[idx])
                if len(norm) >= MIN_DIGITS:
                    numbers.add(norm)
            return numbers
    except (OSError, csv.Error, StopIteration):
        return None


def extract(path: Path) -> tuple[set[str], date | None]:
    """Return (candidate numbers, candidate date) for an incoming file."""
    path = Path(path)
    name = path.name

    found_date = extract_date_from_text(name)
    numbers = extract_numbers_from_text(_strip_dates(name))

    a_number_values = _a_number_column_values(path)
    if a_number_values is not None:
        numbers |= a_number_values
        if found_date is None:
            found_date = extract_date_from_text(_read_content(path))
        return numbers, found_date

    if not numbers or found_date is None:
        content = _read_content(path)
        if found_date is None:
            found_date = extract_date_from_text(content)
        numbers |= extract_numbers_from_text(_strip_dates(content))

    return numbers, found_date
