"""Split a CDR-style CSV (one with an "A Number" column) into one .xlsx file
per unique A Number, so a request only ever receives the rows about its own
number - never the raw multi-number file.

`split_by_column` generalises this to any identifying column and to .xlsx
inputs, so a bulk/combined operator reply can be filtered per record - by IMEI
(Zong / Ufone IMEI) or by A Number (Ufone CDR) - the same way.
"""

from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .parsers import normalize_number
from .staging import staging_dir


def split_by_a_number(path: Path) -> list[tuple[Path, str]] | None:
    """Returns [(per-number .xlsx path, A Number), ...] written into
    main/.staging/ (see staging.py), or None if `path` isn't this CDR format
    (no "A Number" column) - caller should fall back to routing the original
    file as-is.
    """
    if path.suffix.lower() != ".csv":
        return None

    with path.open(newline="", encoding="utf-8-sig", errors="ignore") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return None
        if "A Number" not in header:
            return None

        idx = header.index("A Number")
        groups: dict[str, list[list[str]]] = defaultdict(list)
        for row in reader:
            if len(row) > idx:
                groups[row[idx]].append(row)

    if not groups:
        return None

    out_dir = staging_dir()

    written: list[tuple[Path, str]] = []
    for number, rows in groups.items():
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(header)
        for row in rows:
            ws.append(row)
        out_path = out_dir / f"{path.stem}_{number}.xlsx"
        wb.save(out_path)
        written.append((out_path, number))

    return written


def _read_table(path: Path) -> tuple[list, list[list]] | None:
    """(header, data rows) from a .csv or .xlsx; None if unreadable/empty."""
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            with path.open(newline="", encoding="utf-8-sig", errors="ignore") as f:
                rows = list(csv.reader(f))
            return (rows[0], rows[1:]) if rows else None
        if suffix in {".xlsx", ".xlsm"}:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            try:
                it = wb.active.iter_rows(values_only=True)
                header = next(it, None)
                if header is None:
                    return None
                return list(header), [list(row) for row in it]
            finally:
                wb.close()
    except Exception:
        return None
    return None


def _column_index(header: list, column_names) -> int | None:
    wanted = {n.strip().lower() for n in column_names}
    for i, name in enumerate(header):
        if name is not None and str(name).strip().lower() in wanted:
            return i
    return None


def split_by_column(path: Path, column_names) -> list[tuple[Path, str]] | None:
    """Split a .csv/.xlsx into one .xlsx per distinct normalized value of the
    first column matching any of `column_names` (case-insensitive), written into
    main/.staging/.

    Returns [(per-value .xlsx path, normalized value), ...], or None if the file
    is unreadable or carries none of the requested columns - the caller then
    falls back to routing the file as-is.
    """
    path = Path(path)
    parsed = _read_table(path)
    if parsed is None:
        return None
    header, rows = parsed
    idx = _column_index(header, column_names)
    if idx is None:
        return None

    groups: dict[str, list[list]] = defaultdict(list)
    for row in rows:
        if len(row) <= idx:
            continue
        key = normalize_number(str(row[idx]) if row[idx] is not None else "")
        if key:
            groups[key].append(row)

    if not groups:
        return None

    out_dir = staging_dir()
    written: list[tuple[Path, str]] = []
    for value, value_rows in groups.items():
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(list(header))
        for row in value_rows:
            ws.append(list(row))
        out_path = out_dir / f"{path.stem}_{value}.xlsx"
        wb.save(out_path)
        written.append((out_path, value))
    return written
