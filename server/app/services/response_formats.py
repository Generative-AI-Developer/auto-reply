"""Identify which operator produced an incoming IMEI response from its headers.

A single entered IMEI fans out into four independently-tracked records (one per
operator). When an operator replies, its file is recognised by a small set of
marker columns in the header row, so only that operator's record flips to Sent
- the other three keep waiting for their own reply. This runs before the CDR
`split_by_a_number` step because the Ufone IMEI reply also carries an
"A Number" column and would otherwise be mistaken for a CDR file to split.
"""

from __future__ import annotations

import csv
from pathlib import Path

from .formats import MOBILINK, TELENOR, UFONE, ZONG

# Each operator's IMEI reply is recognised when all of its marker columns are
# present (case-insensitive, whitespace-trimmed). Ufone/Mobilink are checked
# first because their wide layouts are the most distinctive; Telenor/Zong use a
# unique timestamp column so they can't collide with the others.
_SIGNATURES: list[tuple[str, frozenset[str]]] = [
    (UFONE, frozenset({"service provider", "cell sector"})),
    (MOBILINK, frozenset({"calltype", "aparty", "bparty"})),
    (TELENOR, frozenset({"msisdn", "call_start_dt_tm"})),
    (ZONG, frozenset({"mobile no", "last_activity_date"})),
]


def _header_tokens(path: Path) -> set[str]:
    """The first row's non-empty cells, lower-cased and trimmed."""
    suffix = path.suffix.lower()
    try:
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True, data_only=True)
            try:
                cells = next(wb.active.iter_rows(min_row=1, max_row=1, values_only=True), ())
            finally:
                wb.close()
        elif suffix in {".csv", ".txt"}:
            with path.open(newline="", encoding="utf-8-sig", errors="ignore") as f:
                first_line = f.readline()
            # Operator dumps are comma- or tab-separated; pick whichever the
            # header row uses more of.
            delimiter = "\t" if first_line.count("\t") > first_line.count(",") else ","
            cells = next(csv.reader([first_line], delimiter=delimiter), [])
        else:
            return set()
    except Exception:
        # Best-effort: an unreadable/locked file simply isn't recognised.
        return set()
    return {str(c).strip().lower() for c in (cells or ()) if c is not None and str(c).strip()}


def detect_imei_operator(path: Path) -> str | None:
    """Return the network (telenor/mobilink/ufone/zong) whose IMEI-response
    header matches, or None when the file isn't a recognised IMEI response."""
    tokens = _header_tokens(Path(path))
    if not tokens:
        return None
    for network, markers in _SIGNATURES:
        if markers <= tokens:
            return network
    return None


def _digits(s: str) -> str:
    return "".join(ch for ch in s if ch.isdigit())


def imei_column_values(path: Path) -> set[str]:
    """Digit-normalised values from the reply's IMEI column.

    Needed because the generic number extractor short-circuits on the Ufone
    reply's "A Number" column and would otherwise miss the IMEI itself - the
    value the fanned-out per-operator records are keyed on. All four formats
    label the column "IMEI"/"Imei".
    """
    path = Path(path)
    suffix = path.suffix.lower()
    values: set[str] = set()
    try:
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True, data_only=True)
            try:
                rows = wb.active.iter_rows(values_only=True)
                header = next(rows, None) or ()
                idx = _imei_index([str(c).strip().lower() if c is not None else "" for c in header])
                if idx is not None:
                    for row in rows:
                        if idx < len(row) and row[idx] is not None:
                            v = _digits(str(row[idx]))
                            if v:
                                values.add(v)
            finally:
                wb.close()
        elif suffix in {".csv", ".txt"}:
            with path.open(newline="", encoding="utf-8-sig", errors="ignore") as f:
                sample = f.readline()
                delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
                f.seek(0)
                reader = csv.reader(f, delimiter=delimiter)
                header = next(reader, [])
                idx = _imei_index([str(c).strip().lower() for c in header])
                if idx is not None:
                    for row in reader:
                        if idx < len(row) and row[idx]:
                            v = _digits(row[idx])
                            if v:
                                values.add(v)
    except Exception:
        return values
    return values


def _imei_index(header_lc: list[str]) -> int | None:
    for i, name in enumerate(header_lc):
        if name == "imei":
            return i
    return None
